import numpy as np
from openpyxl import load_workbook
from pymoo.core.problem import Problem
from pymoo.optimize import minimize
from pymoo.algorithms.moo.nsga2 import NSGA2
# from pymoo.problems import get_problem
# from pymoo.algorithms.soo.nonconvex.ga import GA

wb = load_workbook("rnd.xlsx", read_only=True)
ws = wb.active

# xlsx non-derived matrix grabber (should do it only for samples)
def load(raw):
    arr = np.empty(np.array(raw).shape)
    for i in range(0, np.array(raw).shape[0]):
        for j in range(0, np.array(raw).shape[1]):
            arr[i][j] = raw[i][j].value
    return arr

# grab sample thetas
theta_CAPM = load(ws['a36':'a45'])
theta_FF3 = load(ws['b36':'b45'])
theta_MACRO = load(ws['c36':'c45'])

# grab rf from spreadsheet
rf = ws['b30'].value

# grab R sample factors
R_i_CAPM = theta_CAPM + rf
R_i_FF3 = theta_FF3 + rf
R_i_MACRO = theta_MACRO + rf

# grab sample w_bar... i.e. Market Weight
w_bar = load(ws['f19':'f28'])

# grab sample w... i.e. Port. Weight
w = load(ws['b19':'b28'])


# grab sample factors
CAPM_factor = ws['b3'].value
FF3_factors = load(ws['b6':'d8'])
MACRO_factors = load(ws['b11':'f15'])


# grab sample beta capm, ff3 and macro
Beta_CAPM = load(ws['o2':'o11'])
Beta_FF3 = load(ws['o14':'q23'])
Beta_MACRO = load(ws['o26':'s35'])

# grab sample sigma sub epsilon, sub t
VC_CAPM_idio = load(ws['ac2':'al11'])
VC_FF3_idio = load(ws['ac14':'al23'])
VC_MACRO_idio = load(ws['ac26':'al35'])

# grab sample get chol (lightgreen)
chol_FF3_VC = load(ws['h6':'j8'])
chol_MACRO_VC = load(ws['h11':'l15'])

# grab idiosyncratics (have formula outside excel)
chol_vc_capm_idio = load(ws['an2':'aw11'])
chol_vc_ff3_idio = load(ws['an14':'aw23'])
chol_vc_macro_idio = load(ws['an26':'aw35'])

# calculate vol_mkt
Vol_Mkt = np.sqrt(CAPM_factor)

# calculate sigma capm, ff3 and macro
Sigma_CAPM_syst = np.dot(Beta_CAPM, Vol_Mkt)
Sigma_FF3_syst = np.dot(Beta_FF3, chol_FF3_VC)
Sigma_MACRO_syst = np.dot(np.dot(Beta_MACRO, chol_MACRO_VC), Beta_MACRO.T)
#truncated last 5 columns # TODO future observation required
Sigma_MACRO_syst = np.delete(Sigma_MACRO_syst, (5, 6, 7, 8, 9), axis=1)

# =MMULT(MINVERSE(MMULT(V14:X23,TRANSPOSE(V14:X23))+MMULT(AN14:AW23,TRANSPOSE(AN14:AW23))),B36:B45)
def get_eta(sigma, idio, theta):
    res1, res2 = (np.dot(sigma, sigma.T)+np.dot(idio, idio.T), theta)
    return np.dot(np.linalg.inv(res1), res2)

eta_capm = get_eta(Sigma_CAPM_syst, chol_vc_capm_idio, theta_CAPM) 
eta_ff3 = get_eta(Sigma_FF3_syst, chol_vc_ff3_idio, theta_FF3)
eta_macro = get_eta(Sigma_MACRO_syst, chol_vc_macro_idio, theta_MACRO)

# calculating systematics
VC_CAPM_syst = np.dot(np.dot(Beta_CAPM, CAPM_factor), Beta_CAPM.T)
VC_FF3_syst = np.dot(np.dot(Beta_FF3, FF3_factors), Beta_FF3.T)
VC_MACRO_syst = np.dot(np.dot(Beta_MACRO, MACRO_factors), Beta_MACRO.T)

#calculating VC
VC_CAPM = VC_CAPM_syst + VC_CAPM_idio
VC_FF3 = VC_FF3_syst + VC_FF3_idio
VC_MACRO = VC_MACRO_syst + VC_MACRO_idio

# calculating b
b_capm = np.sqrt(np.dot(np.dot(w_bar.T, VC_CAPM), w_bar))
b_ff3_0 = np.sqrt(np.dot(np.dot(w_bar.T, VC_FF3), w_bar))
b_ff3 = np.array([[b_ff3_0[0][0]], [0], [0]])
b_macro_0 = np.sqrt(np.dot(np.dot(w_bar.T, VC_MACRO), w_bar))
b_macro = np.array([[b_macro_0[0][0]], [0], [0], [0], [0]])

# =MMULT(MINVERSE(MMULT(V14:X23,TRANSPOSE(V14:X23))+MMULT(AN14:AW23,TRANSPOSE(AN14:AW23))),
# MMULT(V14:X23,BA14:BA16))
def get_pi(sigma, idio, beta):
    res1, res2 = (np.dot(sigma, sigma.T)+np.dot(idio, idio.T), np.dot(sigma,beta))
    return np.dot(np.linalg.inv(res1), res2)

pi_capm = get_pi(Sigma_CAPM_syst, chol_vc_capm_idio, b_capm)
pi_ff3 = get_pi(Sigma_FF3_syst, chol_vc_ff3_idio, b_ff3)
pi_macro = get_pi(Sigma_MACRO_syst, chol_vc_macro_idio, b_macro)

# calculate tau
def get_tau(R_i, eta, pi):
    tau_s1 = R_i + rf
    tau_sumproduct = sum(x*y for x, y in zip(tau_s1, w_bar))
    tau_s2 = -1/2*np.dot(R_i.T, eta)
    tau_s3 = -np.dot(R_i.T, pi)
    return(tau_sumproduct + tau_s2 + tau_s3)

tau_capm = get_tau(R_i_CAPM, eta_capm, pi_capm)
tau_ff3 = get_tau(R_i_FF3, eta_ff3, pi_ff3)
tau_macro = get_tau(R_i_MACRO, eta_macro, pi_macro)

# calculating R
def get_R(R_i):
    return sum(x*y for x, y in zip(w, R_i))

R_CAPM = get_R(R_i_CAPM)
R_FF3 = get_R(R_i_FF3)
R_MACRO = get_R(R_i_MACRO)

# calculating M
def get_M(R_i):
    return sum(x*y for x, y in zip(w_bar, R_i))

M_CAPM = get_M(R_i_CAPM)
M_FF3 = get_M(R_i_FF3)
M_MACRO = get_M(R_i_MACRO)

# '=(EXP(-1/2*MMULT(TRANSPOSE($A$36:$A$45),$AY$2:$AY$11))*1/($B$31*$D$46)+$D$47/$D$46*EXP($BB$2)-1)*AY2'
# calculate growth port
# lol = np.exp(-1/2*np.dot(theta_CAPM.T, eta_capm)*1/(rf*sumproduct_capm)+(sumproduct_ff3/sumproduct_capm))*np.exp(tau_capm)-1*eta_capm[0]
# print(lol)

#######################################################################################
# TODO optimize below
#######################################################################################

theta = theta_CAPM
tau = tau_capm
eta = eta_capm
pi = pi_capm
R = R_CAPM
M = M_CAPM
T = 10
e = np.exp
scalar = np.ndarray.item

def x_star(theta):
    sub_lft_exp = np.dot(scalar(np.divide(1,R)),e(np.dot(np.dot(np.dot(-1/2,theta.T),T),eta)))
    sub_mid_exp = np.dot(np.divide(M,R),e(np.dot(T, tau)))
    lft_exp = np.dot((scalar(np.subtract(np.add(sub_lft_exp,sub_mid_exp),1))),eta)
    rgt_exp = np.dot(scalar(sub_mid_exp),pi)
    return np.add(lft_exp,rgt_exp)

class mproblem(Problem):
    def _evaluate(self, designs, out, *args, **kwargs):
        _ = args
        _ = kwargs
        res = []
        for design in designs:
            # res.append(benchmarks.kursawe(design))
            res.append(x_star(design))
        out["F"] = np.array(res)

# from deap import benchmarks
# benchmarks.kursawe([1,1])
start = [-1, -1, -1, -1, -1, -1, -1, -1, -1, -1]
end = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
# problem = mproblem(n_var=2,n_obj=2,xl=[-5, -5],xu=[5, 5])
problem = mproblem(n_var=10, n_obj=10, xl=start, xu=end)
algorithm = NSGA2(pop_size=3)
stop_criteria = ('n_gen', 100)
res = minimize(problem=problem, algorithm=algorithm, termination=stop_criteria)
print(res.X)
print("----")
print(res.F)

# print(w)
# lol = np.array([0.66297468,  0.15241581,  0.26453545,  0.46613578,  0.03983345,  0.51678012, 0.58905281, -0.43315597, -0.92696819,  0.69194314])
# print(x_star(lol))
# print(w)

# problem = get_problem("g1")
# algorithm = GA(pop_size=10, eliminate_duplicates=True)
# res = minimize(problem, algorithm, seed=1, verbose=False)
# print("X = %s\nF = %s" % (res.X, res.F))
# print(w)

# import plotly.graph_objects as go
# fig=go.Figure(data=go.Scatter(x=res_data))
