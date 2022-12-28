import numpy as np
import copy as cp
print("loading files")
from openpyxl import load_workbook
wb = load_workbook(r"rnd.xlsx")
ws = wb.active
wb_w = load_workbook(r"w.xlsx")  # accessing the weights file
ws_w = wb_w.active  # get the first sheet
print("w.xlsx loaded")
wb_crsp = load_workbook(r"weights_CRSP.xlsx")
ws_crsp = wb_crsp.active  # get the first sheet
print("weights_CRSP.xlsx loaded")
bcapm = load_workbook(r"kalman_betas_camp.xlsx")
ws_bcapm = bcapm.active  # get the first sheet
print("kalman_betas_capm.xlsx loaded")
bff3 = load_workbook(r"betas_ff3.xlsx")
ws_bff3 = bff3.active  # get the first sheet
print("betas_ff3.xlsx loaded")
macro = load_workbook(r"betas_macro.xlsx")
ws_bmacro = macro.active  # get the first sheet
print("betas_macro.xlsx loaded")
capm_factor = load_workbook(r"mgarch_var_cov_capm.xlsx")
ws_capm_factor = capm_factor.active  # get the first sheet
print("mgarch_var_cov_capm.xlsx loaded")
ff3_factor = load_workbook(r"mgarch_var_cov_ff3.xlsx")
ws_ff3_factor = ff3_factor.active  # get the first sheet
print("mgarch_var_cov_ff3.xlsx loaded")
macro_factor = load_workbook(r"mgarch_var_cov_macro.xlsx")
ws_macro_factor = macro_factor.active
print("mgarch_var_cov_macro.xlsx loaded")
capm_idio = load_workbook(r"capm.xlsx")
ws_capm_idio = capm_idio.active
print("capm.xlsx loaded")
ff3_idio = load_workbook(r"ff3.xlsx")
ws_ff3_idio = ff3_idio.active
print("ff3.xlsx loaded")
macro_idio = load_workbook(r"macro.xlsx")
ws_macro_idio = macro_idio.active
print("macro.xlsx loaded")

T = 1
print("T =", T)
my_date = "2014m12"
print("start date =", my_date)
fund_no = "63014"
print("fund_no =", fund_no)
# mode mode mode
global CAPM
CAPM = "CAPM"
global FF3
FF3 = "FF3"
global MACRO
MACRO = "MACRO"
global MODE

MODE = MACRO
print("MODE =", MODE)
# mode mode mode

def set_date():
    global date, date_adj
    date_adj = "'" + date + "'"

time_w_bar = None#[date]
time_w = None#[date_adj]
rf = None#0
#
#
# T - summation. Obtaining incremental dates
def increment_time(date, T):
    all_dates_1 = []
    all_dates_2 = []
    month = int(date[5:])
    for t in range(1, T + 1):
        t = 1
        month = month + t
        if month > 12:
            new_date = str(int(date[0:4]) + 1) + "m"
            new_month = month % 12
            if new_month == 0:
                new_month = 12
            new_date = new_date + str(new_month)
            all_dates_2.append(new_date)
        else:
            new_date_1 = str(int(date[0:4])) + "m"
            new_date_1 = new_date_1 + str(month)
            all_dates_1.append(new_date_1)
    dates_1 = all_dates_1
    dates_2 = all_dates_2
    dates_1.extend(dates_2)
    return (dates_1)

def load(raw):
    arr = np.empty(np.array(raw).shape)
    for i in range(0, np.array(raw).shape[0]):
        for j in range(0, np.array(raw).shape[1]):
            arr[i][j] = raw[i][j].value
    return arr

theta_CAPM = load(ws['a36':'a45'])
theta_FF3 = load(ws['b36':'b45'])
theta_MACRO = load(ws['c36':'c45'])

rf = None
#R_i_CAPM = None
R_i_FF3 = None
R_i_MACRO = None

def grab_w_bar(date):
    all_w = []
    # if T == 2:
    count = 0
    for t in date:
        all_rows = []
        w = []
        [all_rows.append(r) for r in ws_crsp.iter_rows(values_only=True)]
        for i in all_rows:
            if i[0] == t:
                w.append(i[3])
        all_w.append(w)
    return np.array(all_w).T

def grab_w_bar_list(time_w_bar_list):
    MAT = []
    for time_w_bar in time_w_bar_list:
        MAT.append(grab_w_bar([time_w_bar]))
    MAT = np.array(MAT)
    return MAT

def grab_w(date): # so we to make this quarterly
    all_w = []
    # if T == 2:

    for i, t in enumerate(date):
        # creating algorithm in linux will paste here. ok ok.
        # so we have months 3, 6, 9, 12 available for w
        # 1,2,3 -> 3; 4,5,6 ->6; 7,8,9 -> 9; 10,11,12 -> 12. yeah

        t_new = None
        if len(t) == 8:
            num2 = 0
            num = int(t[len(t) - 2])
            if (num % 3):
                num2 = (3 - num % 3)
            n = len(t) - 2
            t_new = t[0:n] + str(num + num2) + "'"
        else:
            num2 = 0
            n = len(t) - 3
            m = len(t) - 1
            num = int(t[n:m])
            if (num % 3):
                num2 = 3 - num % 3
            t_new = t[0:n] + str(num + num2) + "'"

        all_rows = []
        w = []
        [all_rows.append(r) for r in ws_w.iter_rows(values_only=True)]
        for i in all_rows:
            if str(i[0]) == fund_no and i[2] == t_new:
                w.append(i[1])
        all_w.append(w)
    return np.array(all_w).T

def grab_capm_factor(date):
    all_values = []
    for t in date:
        all_rows = []
        [all_rows.append(r) for r in ws_capm_factor.iter_rows(values_only=True)]
        for i in all_rows:
            if i[0] == t:
                all_values.append(i[1])
    return all_values[0]

def grab_capm_factor_list(dates):
    MAT = []
    for date in dates:
        MAT.append(grab_capm_factor([date]))
    MAT = np.array(MAT)
    return MAT

def grab_ff3_factor(date):
    all_values = []
    for t in date:
        all_rows = []
        [all_rows.append(r) for r in ws_ff3_factor.iter_rows(values_only=True)]
        for i in all_rows:
            if i[0] == t:
                all_values.append(i[1:len(i)])
    return ult(list(all_values[0]), (3,3))

def grab_ff3_factor_list(time_w_bar_list):
    ff3_factor_list = []
    for time_w_bar in time_w_bar_list:
        ff3_factor_list.append(grab_ff3_factor([time_w_bar]))
    return np.array(ff3_factor_list)

def grab_macro_factor(date):
    all_values = []
    for t in date:
        all_rows = []
        [all_rows.append(r) for r in ws_macro_factor.iter_rows(values_only=True)]
        for i in all_rows:
            if i[0] == t:
                all_values.append(i[1:len(i)])
    return ult(list(all_values[0]), (5,5))


def grab_macro_factor_list(time_w_bar_list):
    macro_factor_list = []
    for time_w_bar in time_w_bar_list:
        macro_factor_list.append(grab_macro_factor([time_w_bar]))
    return np.array(macro_factor_list)

def grab_beta_capm(date):
    all_betas = []
    for t in date:
        all_rows = []
        [all_rows.append(r) for r in ws_bcapm.iter_rows(values_only=True)]
        for i in all_rows:
            if i[0] == t:
                for x in list(i[1:11]):
                    all_betas.append(x)
    return np.array([all_betas]).T

def grab_beta_capm_list(dates):
    MAT = []
    for date in dates:
        MAT.append(grab_beta_capm([date]))
    MAT = np.array(MAT)
    return MAT

def grab_beta_ff3(date):
    all_betas = []
    for t in date:
        all_rows = []
        [all_rows.append(r) for r in ws_bff3.iter_rows(values_only=True)]
        for i in all_rows:
            if i[0] == t:
                for x in list(i[1:31]):
                    all_betas.append(x)
    arr = np.array([all_betas])
    mat_10_3 = np.zeros((10,3))

    j = 0
    k = 0

    for i in range(len(arr[0])):
        mat_10_3[k][j] = arr[0][i]
        if len(arr[0])/3 > k+1:
            k = k + 1
        else:
            j = j + 1
            k = 0

    return mat_10_3

def grab_beta_ff3_list(dates):
    MAT = []
    for date in dates:
        MAT.append(grab_beta_ff3([date]))
    MAT = np.array(MAT)
    return MAT

def grab_beta_macro(date):
    all_betas = []
    for t in date:
        all_rows = []
        [all_rows.append(r) for r in ws_bmacro.iter_rows(values_only=True)]
        for i in all_rows:
            if i[0] == t:
                for x in list(i[1:51]):
                    all_betas.append(x)
    arr = np.array([all_betas])
    mat_10_5 = np.zeros((10,5))

    j = 0
    k = 0

    for i in range(len(arr[0])):
        mat_10_5[k][j] = arr[0][i]
        if len(arr[0])/5 > k+1:
            k = k + 1
        else:
            j = j + 1
            k = 0

    return mat_10_5

def grab_beta_macro_list(dates):
    MAT = []
    for date in dates:
        MAT.append(grab_beta_macro([date]))
    MAT = np.array(MAT)
    return MAT

def insert_dates_excel(CAPM_start, CAPM_start_year, CAPM_end_year, start_month, end_month, CAPM_end, sheet, t):
    all_months = []
    all_dates = []
    count = 0
    for x in range(CAPM_start_year, CAPM_end_year + 1):
        CAPM_start_new = CAPM_start_year + count
        CAPM_start_new = str(CAPM_start_new) + str("m")
        count += 1
        z = 0
        for i in range(1, 13):
            if CAPM_start_new[0:4] == CAPM_start[0:4]:
                if i >= int(start_month):
                    CAPM_start_new = CAPM_start_new[0:5] + str(i)
                    all_dates.append(CAPM_start_new)
            else:
                CAPM_start_new = CAPM_start_new[0:5] + str(i)
                all_dates.append(CAPM_start_new)
            z += 1
    if global_grab_i == 1 and doit:
        sheet.insert_cols(1)
    all_rows = []
    [all_rows.append(r) for r in sheet.iter_rows(values_only=True)]
    count = 0
    for row in all_rows:
        row = list(row)
        if count >= 1:
            row[0] = all_dates[count - 1]
        count += 1
        if count >= 2:
            if row[0] == t:
                return (np.array(row[1:]))

doit = True 
def insert_dates_excel_list(start_capm, start_year_capm, end_year_capm, start_month_capm, end_month_capm, end_year_str_capm, ws_capm_idio, time_row):
    global doit
    doit = False
    arr = []
    for date in time_row:
        arr.append(insert_dates_excel(start_capm, start_year_capm, end_year_capm, start_month_capm, end_month_capm, end_year_str_capm, ws_capm_idio, date))
    return np.array(arr)

# #---capturing idios---------------------------------
start_capm = "1975m3"
start_year_capm = 1975
end_year_capm = 2016
start_month_capm = "3"
end_month_capm = "12"
end_year_str_capm = "2016m3"
# ult_input = insert_dates_excel(start, start_year, end_year, start_month, end_month, end_year_str, ws_capm_idio, date)
# VC_CAPM_idio = load(ws['ac2':'al11'])-----------------------
shape = (10, 10)

def ult(arr, shape):
    triangle_vector_macro = arr
    red_triangle_macro = np.empty(shape)
    k = 0
    for i in range(shape[0]):
        for j in range(i, shape[0]):
            red_triangle_macro[i][j] = triangle_vector_macro[k]
            k = k + 1
    red_macro = np.empty(shape)
    for i in range(shape[0]):
        for j in range(shape[0]):
            if (j >= i):
                red_macro[i][j] = red_triangle_macro[i][j]
            else:
                red_macro[i][j] = red_triangle_macro.T[i][j]
    return red_macro

# VC_CAPM_idio_list = ult_list(ult_input_capm_list, shape)
def ult_list(ult_input_capm_list, shape):
    red_macro_list = []
    for ult_input_capm in ult_input_capm_list:
        red_macro_list.append(ult(ult_input_capm, shape))
    return np.array(red_macro_list)

# VC_CAPM_idio = ult(ult_input, shape)
# # VC_FF3_idio = load(ws['ac14':'al23'])----------------------------------
start_ff3 = "1975m3"
start_year_ff3 = 1975
end_year_ff3 = 2016
start_month_ff3 = "3"
end_month_ff3 = "12"
end_year_str_ff3 = "2016m3"
# ult_input = insert_dates_excel(start, start_year, end_year, start_month, end_month, end_year_str, ws_ff3_idio, date)
shape = (10, 10)

# VC_MACRO_idio = load(ws['ac26':'al35'])---------------------------------
start_macro = "1987m8"
start_year_macro = 1987
end_year_macro = 2016
start_month_macro = "8"
end_month_macro = "12"
end_year_str_macro = "2016m3"
# ult_input = insert_dates_excel(start, start_year, end_year, start_month, end_month, end_year_str, ws_macro_idio, date)
# VC_CAPM_idio = load(ws['ac2':'al11'])-----------------------
shape = (10, 10)

# =MMULT(MINVERSE(MMULT(V14:X23,TRANSPOSE(V14:X23))+MMULT(AN14:AW23,TRANSPOSE(AN14:AW23))),B36:B45)
def get_eta(sigma, idio, theta):
    res1, res2 = (np.dot(sigma, sigma.T) + np.dot(idio, idio.T), theta)
    return np.dot(np.linalg.inv(res1), res2)

def get_eta_list(sigma_list, idio_list, theta):
    eta_list = []
    for i in range(len(sigma_list)):
        eta_list.append(get_eta(sigma_list[i], idio_list[i], theta))
    return np.array(eta_list)

# =MMULT(MINVERSE(MMULT(V14:X23,TRANSPOSE(V14:X23))+MMULT(AN14:AW23,TRANSPOSE(AN14:AW23))),
# MMULT(V14:X23,BA14:BA16))
def get_pi(sigma, idio, beta):
    res1, res2 = (np.dot(sigma, sigma.T) + np.dot(idio, idio.T), np.dot(sigma, beta))
    return np.dot(np.linalg.inv(res1), res2)

def get_pi_list(sigma_list, idio_list, beta_list):
    pi_list = []
    for i in range(len(sigma_list)):
        pi_list.append(get_pi(sigma_list[i], idio_list[i], beta_list[i]))
    return np.array(pi_list)

# calculate tau
def get_tau(R_i, eta, pi,w_bar):
    tau_s1 = R_i + rf

    tau_sumproduct = sum(x * y for x, y in zip(tau_s1, w_bar))
    tau_s2 = -1 / 2 * np.dot(R_i.T, eta)
    tau_s3 = -np.dot(R_i.T, pi)
    return (tau_sumproduct + tau_s2 + tau_s3)

def get_tau_list(R_i_CAPM, eta_capm_list, pi_capm_list, w_bar_list):
    tau_list = []
    for i in range(len(eta_capm_list)):
        tau_list.append(get_tau(R_i_CAPM, eta_capm_list[i], pi_capm_list[i], w_bar_list[i]))
    return np.array(tau_list)

# calculating R
def get_R(R_i,w):
    return sum(x * y for x, y in zip(w, R_i))

# calculating M
def get_M(R_i,w_bar):
    return sum(x * y for x, y in zip(w_bar, R_i))

#theta = theta_CAPM
tau = None#tau_capm
eta = 0 #why is this 0?

# our major malfunctions these need to be dealt in the globals
pi = None#pi_capm
R = None#R_CAPM
M = None#M_CAPM
e = np.exp
scalar = np.ndarray.item

date_index = 0
time_list = np.array(increment_time(my_date, T))

time_row = []
time_row.append(my_date)
for i, t in enumerate(time_list):
    #time_row[i] = time_list[i]
    if t != time_list[len(time_list)-1]:
        time_row.append(t)

global_grab_i = 0
def global_grab(theta):
    global global_grab_i
    global_grab_i = global_grab_i + 1

    global R_T_1
    global M_T_1
    global eta_T_1
    global pi_T_1

    global set_once
    if set_once:
        global w_bar # required
        global w_bar_list # required
        global w # required
        global chol_vc_capm_idio # required
        global chol_vc_capm_idio_list # required
        global chol_vc_ff3_idio # required
        global chol_vc_ff3_idio_list # required
        global chol_vc_macro_idio # required
        global chol_vc_macro_idio_list # required
        global Sigma_CAPM_syst # required
        global Sigma_CAPM_syst_list # required
        global Sigma_FF3_syst # required
        global Sigma_FF3_syst_list # required
        global Sigma_MACRO_syst # required
        global Sigma_MACRO_syst_list # required
        global pi_capm # required
        global pi_capm_list # required
        global pi_ff3 # required
        global pi_ff3_list # required
        global pi_macro # required
        global pi_macro_list # required
        global pi # required
        global w_T_1

        global date
        global time_row
        date = time_row[0]
        set_date()

        time_w_bar = [date]
        time_w = [date_adj]

        time_w_list = []
        for i in range(len(time_row)):
            date_str = time_row[i]
            date_adj_str =  "'" + date_str + "'"
            time_w_list.append(date_adj_str)

        time_w_bar_list = time_row

        w_bar = grab_w_bar(time_w_bar)
        w_bar_list = grab_w_bar_list(time_w_bar_list)

        w = grab_w(time_w)
        w_T_1 = w

        CAPM_factor = grab_capm_factor(time_w_bar)
        CAPM_factor_list = grab_capm_factor_list(time_w_bar_list)
        Vol_Mkt_capm = np.sqrt(CAPM_factor)
        Vol_Mkt_capm_list = []
        for CAPM_factor in CAPM_factor_list:
            Vol_Mkt_capm_list.append(np.sqrt(CAPM_factor))

        FF3_factor = grab_ff3_factor(time_w_bar)
        FF3_factor_list = grab_ff3_factor_list(time_w_bar_list)
        Vol_Mkt_ff3 = (np.linalg.cholesky(FF3_factor)).T
        Vol_Mkt_ff3_list = []
        for FF3_factor in FF3_factor_list:
            Vol_Mkt_ff3_list.append((np.linalg.cholesky(FF3_factor)).T)
        Vol_Mkt_ff3_list = np.array(Vol_Mkt_ff3_list)

        MACRO_factor = grab_macro_factor(time_w_bar)
        MACRO_factor_list = grab_macro_factor_list(time_w_bar_list)
        Vol_Mkt_macro = (np.linalg.cholesky(MACRO_factor)).T
        Vol_Mkt_macro_list = []
        for MACRO_factor in MACRO_factor_list:
            Vol_Mkt_macro_list.append((np.linalg.cholesky(MACRO_factor)).T)
        Vol_Mkt_macro_list = np.array(Vol_Mkt_macro_list)

        Beta_CAPM = grab_beta_capm(time_w)
        Beta_CAPM_list = grab_beta_capm_list(time_w_list)

        Beta_FF3 = grab_beta_ff3(time_w)
        Beta_FF3_list = grab_beta_ff3_list(time_w_list)

        Beta_MACRO = grab_beta_macro(time_w)
        Beta_MACRO_list = grab_beta_macro_list(time_w_list)

        ult_input_capm = insert_dates_excel(start_capm, start_year_capm, end_year_capm, start_month_capm, end_month_capm, end_year_str_capm, ws_capm_idio, date)
        ult_input_ff3 = insert_dates_excel(start_ff3, start_year_ff3, end_year_ff3, start_month_ff3, end_month_ff3, end_year_str_ff3, ws_ff3_idio, date)
        ult_input_macro = insert_dates_excel(start_macro, start_year_macro, end_year_macro, start_month_macro, end_month_macro, end_year_str_macro, ws_macro_idio, date)

        ult_input_capm_list = insert_dates_excel_list(start_capm, start_year_capm, end_year_capm, start_month_capm, end_month_capm, end_year_str_capm, ws_capm_idio, time_row)
        VC_CAPM_idio = ult(ult_input_capm, shape)
        VC_CAPM_idio_list = ult_list(ult_input_capm_list, shape)

        ult_input_ff3_list = insert_dates_excel_list(start_ff3, start_year_ff3, end_year_ff3, start_month_ff3, end_month_ff3, end_year_str_ff3, ws_ff3_idio, time_row)
        VC_FF3_idio = ult(ult_input_ff3, shape)
        VC_FF3_idio_list = ult_list(ult_input_ff3_list, shape)

        ult_input_macro_list = insert_dates_excel_list(start_macro, start_year_macro, end_year_macro, start_month_macro, end_month_macro, end_year_str_macro, ws_macro_idio, time_row)
        VC_MACRO_idio = ult(ult_input_macro, shape)
        VC_MACRO_idio_list = ult_list(ult_input_macro_list, shape)

        chol_vc_capm_idio = np.linalg.cholesky(VC_CAPM_idio).T
        chol_vc_capm_idio_list = []
        for VC_CAPM_idio in VC_CAPM_idio_list:
            chol_vc_capm_idio_list.append(np.linalg.cholesky(VC_CAPM_idio).T)
        chol_vc_capm_idio_list = np.array(chol_vc_capm_idio_list)

        chol_vc_ff3_idio = np.linalg.cholesky(VC_FF3_idio).T
        chol_vc_ff3_idio_list = []
        for VC_FF3_idio in VC_FF3_idio_list:
            chol_vc_ff3_idio_list.append(np.linalg.cholesky(VC_FF3_idio).T)
        chol_vc_ff3_idio_list = np.array(chol_vc_ff3_idio_list)

        chol_vc_macro_idio = np.linalg.cholesky(VC_MACRO_idio).T
        chol_vc_macro_idio_list = []
        for VC_MACRO_idio in VC_MACRO_idio_list:
            chol_vc_macro_idio_list.append(np.linalg.cholesky(VC_MACRO_idio).T)
        chol_vc_macro_idio_list = np.array(chol_vc_macro_idio_list)

        Sigma_CAPM_syst = np.dot(Beta_CAPM, Vol_Mkt_capm)
        Sigma_CAPM_syst_list = []
        for i in range(len(Vol_Mkt_capm_list)):
            Sigma_CAPM_syst_list.append(np.dot(Beta_CAPM_list[i], Vol_Mkt_capm_list[i]))
        Sigma_CAPM_syst_list = np.array(Sigma_CAPM_syst_list)

        Sigma_FF3_syst = np.dot(Beta_FF3, Vol_Mkt_ff3)
        Sigma_FF3_syst_list = []
        for i in range(len(Vol_Mkt_ff3_list)):
            Sigma_FF3_syst_list.append(np.dot(Beta_FF3_list[i], Vol_Mkt_ff3_list[i]))
        Sigma_FF3_syst_list = np.array(Sigma_FF3_syst_list)

        # ask waleed bout this
        Sigma_MACRO_syst = np.dot(np.dot(Beta_MACRO, Vol_Mkt_macro), Beta_MACRO.T)
        Sigma_MACRO_syst = np.delete(Sigma_MACRO_syst, (5, 6, 7, 8, 9), axis=1)

        Sigma_MACRO_syst_list = []
        for i in range(len(Vol_Mkt_macro_list)):
            tmp = np.dot(np.dot(Beta_MACRO_list[i], Vol_Mkt_macro_list[i]), Beta_MACRO_list[i].T)
            tmp = np.delete(tmp, (5, 6, 7, 8, 9), axis=1)
            Sigma_MACRO_syst_list.append(tmp)
        Sigma_MACRO_syst_list = np.array(Sigma_MACRO_syst_list)

        VC_CAPM_syst = np.dot(np.dot(Beta_CAPM, CAPM_factor), Beta_CAPM.T)
        VC_CAPM_syst_list = []
        for i in range(len(Beta_CAPM_list)):
            # VC_CAPM_syst_list.append(np.dot(np.dot(Beta_CAPM_list[i], CAPM_factor_list[(len(CAPM_factor_list)-1)-i]), Beta_CAPM_list[i].T))
            VC_CAPM_syst_list.append(np.dot(np.dot(Beta_CAPM_list[i], CAPM_factor_list[i]), Beta_CAPM_list[i].T))
        VC_CAPM_syst_list = np.array(VC_CAPM_syst_list)
        VC_CAPM = VC_CAPM_syst + VC_CAPM_idio
        VC_CAPM_list = []
        for i in range(len(VC_CAPM_syst_list)):
            VC_CAPM_list.append(VC_CAPM_syst_list[i] + VC_CAPM_idio_list[i])
        VC_CAPM_list = np.array(VC_CAPM_list)
        b_capm = np.sqrt(np.dot(np.dot(w_bar.T, VC_CAPM), w_bar))
        b_capm_list = []
        for i in range(len(w_bar_list)):
            b_capm_list.append(np.sqrt(np.dot(np.dot(w_bar_list[i].T, VC_CAPM_list[i]), w_bar_list[i])))
        b_capm_list = np.array(b_capm_list)
        pi_capm = get_pi(Sigma_CAPM_syst, chol_vc_capm_idio, b_capm)
        pi_capm_list = get_pi_list(Sigma_CAPM_syst_list, chol_vc_capm_idio_list, b_capm_list)

        VC_FF3_syst = np.dot(np.dot(Beta_FF3, FF3_factor), Beta_FF3.T)
        VC_FF3_syst_list = []
        for i in range(len(Beta_FF3_list)):
            # VC_FF3_syst_list.append(np.dot(np.dot(Beta_FF3_list[i], FF3_factor_list[(len(FF3_factor_list)-1)-i]), Beta_FF3_list[i].T))
            VC_FF3_syst_list.append(np.dot(np.dot(Beta_FF3_list[i], FF3_factor_list[i]), Beta_FF3_list[i].T))
        VC_FF3_syst_list = np.array(VC_FF3_syst_list)
        VC_FF3 = VC_FF3_syst + VC_FF3_idio
        VC_FF3_list = []
        for i in range(len(VC_FF3_syst_list)):
            VC_FF3_list.append(VC_FF3_syst_list[i] + VC_FF3_idio_list[i])
        VC_FF3_list = np.array(VC_FF3_list)
        b_ff3 = np.sqrt(np.dot(np.dot(w_bar.T, VC_FF3), w_bar))
        zero_matrix = np.zeros(b_ff3.shape)
        b_ff3 = np.vstack([b_ff3, zero_matrix, zero_matrix])
        b_ff3_list = []
        for i in range(len(w_bar_list)):
            tmp_ff3 = np.sqrt(np.dot(np.dot(w_bar_list[i].T, VC_FF3_list[i]), w_bar_list[i]))
            tmp_ff3 = np.vstack([tmp_ff3, zero_matrix, zero_matrix])
            b_ff3_list.append(cp.copy(tmp_ff3))
        b_ff3_list = np.array(b_ff3_list)
        pi_ff3 = get_pi(Sigma_FF3_syst, chol_vc_ff3_idio, b_ff3)
        pi_ff3_list = get_pi_list(Sigma_FF3_syst_list, chol_vc_ff3_idio_list, b_ff3_list)
        # b_ff3_list x

        VC_MACRO_syst = np.dot(np.dot(Beta_MACRO, MACRO_factor), Beta_MACRO.T)
        VC_MACRO_syst_list = []
        for i in range(len(Beta_MACRO_list)):
            # VC_MACRO_syst_list.append(np.dot(np.dot(Beta_MACRO_list[i], MACRO_factor_list[(len(MACRO_factor_list)-1)-i]), Beta_MACRO_list[i].T))
            VC_MACRO_syst_list.append(np.dot(np.dot(Beta_MACRO_list[i], MACRO_factor_list[i]), Beta_MACRO_list[i].T))
            # MACRO_factor_list ok; Beta_MACRO_list ok;
        VC_MACRO_syst_list = np.array(VC_MACRO_syst_list)
        VC_MACRO = VC_MACRO_syst + VC_MACRO_idio
        VC_MACRO_list = []
        for i in range(len(VC_MACRO_syst_list)):
            VC_MACRO_list.append(VC_MACRO_syst_list[i] + VC_MACRO_idio_list[i])
        VC_MACRO_list = np.array(VC_MACRO_list)
        b_macro = np.sqrt(np.dot(np.dot(w_bar.T, VC_MACRO), w_bar))
        zero_matrix = np.zeros(b_macro.shape)
        b_macro = np.vstack([b_macro, zero_matrix, zero_matrix, zero_matrix, zero_matrix])
        b_macro_list = []
        for i in range(len(w_bar_list)):
            tmp_macro = np.sqrt(np.dot(np.dot(w_bar_list[i].T, VC_MACRO_list[i]), w_bar_list[i]))
            tmp_macro = np.vstack([tmp_macro, zero_matrix, zero_matrix, zero_matrix, zero_matrix])
            b_macro_list.append(cp.copy(tmp_macro))
        b_macro_list = np.array(b_macro_list)
        pi_macro = get_pi(Sigma_MACRO_syst, chol_vc_macro_idio, b_macro)
        pi_macro_list = get_pi_list(Sigma_MACRO_syst_list, chol_vc_macro_idio_list, b_macro_list)

        set_once = False

    # depending upon theta
    global rf
    rf = 0

    if (MODE == CAPM):
        pi = pi_capm
    if (MODE == FF3):
        pi = pi_ff3
    if (MODE == MACRO):
        pi = pi_macro

    if (MODE == CAPM):
        global R_i_CAPM
        R_i_CAPM = theta + rf
    if (MODE == FF3):
        global R_i_FF3
        R_i_FF3 = theta + rf
    if (MODE == MACRO):
        global R_i_MACRO
        R_i_MACRO = theta + rf

    if MODE == CAPM:
        global eta_capm
        global eta_capm_list
        eta_capm = get_eta(Sigma_CAPM_syst, chol_vc_capm_idio, theta)
        eta_capm_list = get_eta_list(Sigma_CAPM_syst_list, chol_vc_capm_idio_list, theta) 
    if MODE == FF3:
        global eta_ff3
        global eta_ff3_list
        eta_ff3 = get_eta(Sigma_FF3_syst, chol_vc_ff3_idio, theta)
        eta_ff3_list = get_eta_list(Sigma_FF3_syst_list, chol_vc_ff3_idio_list, theta) 
    if MODE == MACRO:
        global eta_macro
        global eta_macro_list
        eta_macro = get_eta(Sigma_MACRO_syst, chol_vc_macro_idio, theta)
        eta_macro_list = get_eta_list(Sigma_MACRO_syst_list, chol_vc_macro_idio_list, theta) 

    global eta
    if MODE == CAPM:
        eta = eta_capm
    if MODE == FF3:
        eta = eta_ff3
    if MODE == MACRO:
        eta = eta_macro

    if MODE == CAPM:
        global R_CAPM
        R_CAPM = get_R(R_i_CAPM, w)
    if MODE == FF3:
        global R_FF3
        R_FF3 = get_R(R_i_FF3, w)
    if MODE == MACRO:
        global R_MACRO
        R_MACRO = get_R(R_i_MACRO, w)

    global R
    if MODE == CAPM:
        R = R_CAPM
    if MODE == FF3:
        R = R_FF3
    if MODE == MACRO:
        R = R_MACRO

    if MODE == CAPM:
        global M_CAPM
        M_CAPM = get_M(R_i_CAPM, w_bar)
    if MODE == FF3:
        global M_FF3
        M_FF3 = get_M(R_i_FF3, w_bar)
    if MODE == MACRO:
        global M_MACRO
        M_MACRO = get_M(R_i_MACRO, w_bar)

    global M
    if MODE == CAPM:
        M = M_CAPM
    if MODE == FF3:
        M = M_FF3
    if MODE == MACRO:
        M = M_MACRO

    if MODE == CAPM:
        global tau_capm
        global tau_capm_list
        tau_capm = get_tau(R_i_CAPM, eta_capm, pi_capm, w_bar)
        tau_capm_list = get_tau_list(R_i_CAPM, eta_capm_list, pi_capm_list, w_bar_list)

    if MODE == FF3:
        global tau_ff3
        global tau_ff3_list
        tau_ff3 = get_tau(R_i_FF3, eta_ff3, pi_ff3, w_bar)
        tau_ff3_list = get_tau_list(R_i_FF3, eta_ff3_list, pi_ff3_list, w_bar_list)
        # eta_ff3_list ok; pi_ff3_list x;

    if MODE == MACRO:
        global tau_macro
        global tau_macro_list
        tau_macro = get_tau(R_i_MACRO, eta_macro, pi_macro, w_bar)
        tau_macro_list = get_tau_list(R_i_MACRO, eta_macro_list, pi_macro_list, w_bar_list)

    global tau
    if MODE == CAPM:
        tau = tau_capm
    if MODE == FF3:
        tau = tau_ff3
    if MODE == MACRO:
        tau = tau_macro

    global lv_first_integral
    global lv_second_integral

    if MODE == CAPM:
        lv_first_integral = integral1(theta, eta_capm_list)
        lv_second_integral = integral2(tau_capm_list)
    if MODE == FF3:
        lv_first_integral = integral1(theta, eta_ff3_list)
        lv_second_integral = integral2(tau_ff3_list)
    if MODE == MACRO:
        lv_first_integral = integral1(theta, eta_macro_list)
        lv_second_integral = integral2(tau_macro_list)

    pi_T_1 = pi
    R_T_1 = R
    M_T_1 = M
    eta_T_1 = eta

def integral1(theta, eta_capm_list):
    res = -1/2 * np.dot(theta.T, eta_capm_list[0])
    for i in range(1, len(eta_capm_list)):
        res += -1/2 * np.dot(theta.T, eta_capm_list[i])
    return res

def integral2(tau_capm_list):
    res = cp.copy(tau_capm_list[0])
    for i in range(1, len(tau_capm_list)):
        res += tau_capm_list[i]
    return res

M_T_1 = None
R_T_1 = None
eta_T_1 = None
pi_T_1 = None 
w_T_1 = None
lv_first_integral = None 
lv_second_integral = None 
set_once = True

def x_star(theta):
    global_grab(theta)
    lft_lft_exp = (1 / R_T_1 * e(lv_first_integral))
    # print("lft_lft_exp")
    # print(lft_lft_exp)
    lft_mid_exp = M_T_1 / R_T_1 * e(lv_second_integral)
    # print("lft_mid_exp")
    # print(lft_mid_exp)
    # print("M_T_1")
    # print(M_T_1)
    # print("R_T_1")
    # print(R_T_1)
    # print("tau -> lv_second_integral")
    # print(lv_second_integral)
    # print("lft_mid_exp")
    # print(lft_mid_exp)
    # print("eta_T_1")
    # print(eta_T_1)
    # print("pi_T_1")
    # print(pi_T_1)
    lft_exp = (lft_lft_exp + lft_mid_exp - 1) * eta_T_1
    rgt_exp = lft_mid_exp * pi_T_1
    # print("lft_exp")
    # print(lft_exp)
    # print("rgt_exp")
    # print(rgt_exp)
    # quit()
    # print("x_star(theta_MACRO)")
    return lft_exp + rgt_exp

# for date = 2014m12; T = 1; fund_no = "63014"
# theta_test = np.array([[0.07255292, 0.08408584, 0.06552885, 0.06671856, 0.07544954, 0.0596966, 0.07516593, 0.01340695, 0.04415142, 0.06795616]])
theta_test = np.array([[0.11012354, 0.13017151, 0.10203563, 0.11835238, 0.13837704, 0.1188936, 0.10596325, 0.08390391, 0.07149853, 0.12371767]])
theta_test = theta_test.T
print(x_star(theta_test))
print(np.sum(x_star(theta_test)))


# for date = 2014m1 -> 2014m12; T = 12; fund_no = "63014"
# theta_test2 = np.array([[0.02301814, 0.03238492, 0.02421527, 0.0248188,  0.02617493, 0.02154602, 0.02242076, 0.01509845, 0.01224412, 0.02736836]])
# theta_test2 = theta_test2.T
# print(x_star(theta_test2))

# print(x_star(theta_MACRO))

# [[0.10609662] [0.11991286] [0.001072  ] [0.05679585] [0.04062507] [0.01680532] [0.35979177] [0.16368092] [0.0115218 ] [0.12377111]]

def distance(x_star, w):
    r0 = (x_star[0] - w[0]) ** 2
    r1 = (x_star[1] - w[1]) ** 2
    r2 = (x_star[2] - w[2]) ** 2
    r3 = (x_star[3] - w[3]) ** 2
    r4 = (x_star[4] - w[4]) ** 2
    r5 = (x_star[5] - w[5]) ** 2
    r6 = (x_star[6] - w[6]) ** 2
    r7 = (x_star[7] - w[7]) ** 2
    r8 = (x_star[8] - w[8]) ** 2
    r9 = (x_star[9] - w[9]) ** 2
    return np.sqrt(r0 + r1 + r2 + r3 + r4 + r5 + r6 + r7 + r8 + r9)

# w = None
def cwc(theta):
    global w_T_1
    res = x_star(theta)
    # return np.append(res, distance(res, w_T_1))
    return np.append(res, distance(res, w_T_1))
